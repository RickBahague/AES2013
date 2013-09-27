# Script to clean-up downloaded COMELEC Project of Precincts 
# Raw data should be on stored in Microsoft spreadsheet format 
# Author: Rick Bahague
# Email: rick@cp-union.com / rick@opensourceshoppe.com
# GPLv3 Release


from openpyxl import load_workbook
from openpyxl import Workbook
import sys, os, csv

def check_sheet_format(sheet):

	sheet_data = wb.get_sheet_by_name(sheet)
	if sheet_data.cell('A5').value == '':
		print("Check sheet:" + sheet)

def get_municipal_data(sheet, province, region):

	sheet_data = wb.get_sheet_by_name(sheet)
	sheet_max_row = sheet_data.get_highest_row()
	municipal_data = {'Region':str(region),'Province':str(province),'Municipality': sheet,'Barangays':sheet_data.cell(row = sheet_max_row - 4, column = 3).value,'Precincts':sheet_data.cell(row = sheet_max_row - 3, column = 3).value,'TRV':sheet_data.cell(row = sheet_max_row - 2, column = 3).value,'CPs':sheet_data.cell(row = sheet_max_row - 1, column = 3).value}
	return municipal_data

def get_polling_center_data(sheet,clustered_precincts,province,region):

	sheet_data = wb.get_sheet_by_name(sheet)
	sheet_max_row = sheet_data.get_highest_row()
	data = {'Region': region,'Province': province,'Municipality': sheet,'Barangay':'','Center':'','Cp_no':'','TRV':''}
	sheet_max_row = sheet_data.get_highest_row()

	for i in xrange(0,sheet_max_row):

		temp = sheet_data.cell(row = i, column = 0).value
		if(temp != None and sheet_data.cell(row = i, column = 1).value == None and temp!='BARANGAY'):
			barangay = temp

		temp = sheet_data.cell(row = i, column = 1).value
		if(temp !=None and sheet_data.cell(row = i, column = 2).value == None and data['Center'] == '' and temp!='SUB TOTAL'):
			center = temp

		#temp = sheet_data.cell(row = i, column = 1).value
		#if(temp != None and sheet_data.cell(row = i, column = 2).value == None and temp!='SUB TOTAL'):
		#	data['Location'] = temp

		temp = sheet_data.cell(row = i, column = 5).value
		check = sheet_data.cell(row = i, column = 6).value

		if(temp != None and type(check) is int):
			data['Region'] = region
			data['Province'] = province
			data['Municipality'] = sheet
			data['Barangay'] = barangay
			data['Cp_no'] = sheet_data.cell(row = i, column = 4).value
			data['TRV'] = sheet_data.cell(row = i, column = 6).value
			data['Center'] = center
			clustered_precincts.append(data)
			data = {'Region': region,'Province': province,'Municipality': sheet,'Barangay':data['Barangay'],'Center':'','Cp_no':'','TRV':''}
			#print(clustered_precincts)

		if(sheet_data.cell(row = i, column = 1).value == "SUB TOTAL"):
			data = {'Region': region,'Province': province,'Municipality': sheet,'Barangay':'','Center':'','Cp_no':'','TRV':''}

	return clustered_precincts


def write_municipal_csv(municipal_data,workbook):
	# writer code from http://www.gadzmo.com/python/reading-and-writing-csv-files-with-python-dictreader-and-dictwriter/
	reload(sys)
	sys.setdefaultencoding('utf8')

	print ('creating csv for municipal summary')
	fieldnames = ['Region','Province','Municipality', 'Barangays', 'Precincts', 'TRV', 'CPs']
	test_file = open(workbook + '-municipal.csv','wb')
	test_file_all = open('All Municipal.csv','a')
	csvwriter = csv.DictWriter(test_file, delimiter=',', fieldnames=fieldnames)
	csvwriter_all = csv.DictWriter(test_file_all, delimiter=',', fieldnames=fieldnames)
	csvwriter.writerow(dict((fn,fn) for fn in fieldnames))
	csvwriter_all.writerow(dict((fn,fn) for fn in fieldnames))
	for row in municipal_data:
		csvwriter.writerow(row)
		csvwriter_all.writerow(row)
	test_file.close()
	test_file_all.close()

	reload(sys)


def write_clustered_csv(clustered_precincts,workbook):
	# writer code from http://www.gadzmo.com/python/reading-and-writing-csv-files-with-python-dictreader-and-dictwriter/
	reload(sys)
	sys.setdefaultencoding('utf8')

	print ('creating csv for clustered_precincts')
	fieldnames = ['Region','Province','Municipality', 'Barangay', 'Center', 'Cp_no','TRV']
	test_file = open(workbook + '-clusters.csv','wb')
	test_file_all = open('All Clusters.csv','a')
	csvwriter = csv.DictWriter(test_file, delimiter=',', fieldnames=fieldnames)
	csvwriter_all = csv.DictWriter(test_file_all, delimiter=',', fieldnames=fieldnames)
	csvwriter.writerow(dict((fn,fn) for fn in fieldnames))
	csvwriter_all.writerow(dict((fn,fn) for fn in fieldnames))
	for row in clustered_precincts:
		#csvwriter.writerow([s.encode('unicode') if type(s) is unicode else s for s in row])
		csvwriter.writerow(row)
		csvwriter_all.writerow(row)
	test_file.close()
	test_file_all.close()
	reload(sys)

''' --- Main Program '''

path = os.getcwd()
workbooks = os.listdir(path + '/Raw')

for workbook in workbooks:
	
	region = workbook.split("-",1)[0]
	province_temp = workbook.split("-",1)[1]
	province = province_temp.split(".",1)[0]

	wb = load_workbook(path + '/Raw/' + workbook)
	sheets = wb.get_sheet_names()
	
	municipal_data = []
	clustered_precincts = []

	print ('Workbook processing initiated for ' + workbook)

	for sheet in sheets:
		print ('processing municipal data for: ' + sheet)
		municipal_data.append(get_municipal_data(sheet,province,region))
		print ('processing clustered precincts for: ' + sheet)
		clustered_precincts = get_polling_center_data(sheet,clustered_precincts,province,region)
		#print(clustered_precincts)

	#print(municipal_data)

	write_municipal_csv(municipal_data,workbook.split(".",1)[0])

	write_clustered_csv(clustered_precincts,workbook.split(".",1)[0])

